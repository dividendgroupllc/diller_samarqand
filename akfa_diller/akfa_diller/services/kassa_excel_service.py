import re
from typing import Dict, List, Optional

import frappe
from frappe import _

from akfa_diller.akfa_diller.utils.helpers import get_file_path, parse_numeric

# Fixed column order of the "kassa" export template (confirmed against real
# files from both sample branches -- unlike the Sales/Purchase Excel format,
# this is a rigid single-template export, not free-form, so positional
# columns (not header-text matching per column) is the right approach here.
_COLUMNS_PAYMENTS = (
    "date", "diler", "payment_type", "person_text", "target_text",
    "izoh", "amount", "currency", "creator", "updater",
    "created", "updated", "status",
)

# Second, unrelated export template ("Список получения" / "list of receipts")
# seen from some branches (confirmed 2026-08-08): ID, Payer, Account
# (NAXT/Plastik), Information (free text, NOT a clean field -- sometimes a
# large number, sometimes a note like "vozvorat zol T"; stored as-is in
# izoh), Date, Sum. No Diler/payment_type/Holat columns at all -- unlike
# _COLUMNS_PAYMENTS this format is always exactly one branch per file (the
# Kassa Import doc's own `branch` field supplies what "Diler" would have),
# and has no DELETED-row concept to filter.
_COLUMNS_RECEIPTS_LIST = ("id", "person_text", "payment_account", "info", "date", "amount")

_PHONE_RE = re.compile(r"\+?\d[\d\s]{7,}\d")


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa
    except ImportError:
        frappe.throw(_("openpyxl o'rnatilmagan. Ishga tushiring: pip install openpyxl"))


def _normalize_payment_account(raw: Optional[str]) -> str:
    """"NAXT"/"Naqt"/etc -> "NAQT"; anything else non-empty (Plastik, Karta, ...) -> "PLASTIK".
    Defaults to "NAQT" when the source format has no such column at all."""
    text = (raw or "").strip().upper()
    if not text or "NA" in text[:3]:  # NAXT / NAQT / NAQD all start with "NA"
        return "NAQT"
    return "PLASTIK"


def read_kassa_report(file_url: str) -> List[Dict]:
    """
    Read a "kassa" (cash register) Excel export -- either of the two known
    source templates (see _read_payments_format / _read_receipts_list_format).

    Returns one dict per ACTIVE row, in file order, with keys: row_num, date
    (DD.MM.YYYY string, as-is), diler, payment_type (RECEIPT/PAYMENT),
    person_text, izoh, amount (float, sign preserved), payment_account
    ("NAQT" or "PLASTIK" -- which real cash account this row belongs to).
    """
    _ensure_openpyxl()
    from openpyxl import load_workbook

    file_path = get_file_path(file_url)
    if not file_path:
        frappe.throw(_("Excel fayl topilmadi: {0}").format(file_url))

    wb = load_workbook(file_path, data_only=True)
    try:
        ws = wb["Template"] if "Template" in wb.sheetnames else wb.active
        header_row, fmt = _find_header_row(ws)
        if not header_row:
            frappe.throw(
                _("Excel formati tanilmadi -- na 'Sana'/'Diler' (to'lovlar hisoboti), "
                  "na 'ИД'/'Плательщик' (qabul qilish ro'yxati) sarlavhasi topilmadi. "
                  "Bu 'Остаток' (ombor qoldig'i) fayli bo'lishi mumkin -- Kassa Import "
                  "faqat to'lov/kassa hisobotlarini qabul qiladi.")
            )

        if fmt == "payments":
            return _read_payments_format(ws, header_row)
        return _read_receipts_list_format(ws, header_row)
    finally:
        wb.close()


def _read_payments_format(worksheet, header_row: int) -> List[Dict]:
    """`DELETED` rows are dropped entirely here -- they are not real
    transactions (confirmed against real data: some show up mid-file with
    every other field identical to an active row, just re-flagged)."""
    rows = []
    for row_num, row in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, max_col=len(_COLUMNS_PAYMENTS), values_only=True),
        start=header_row + 1,
    ):
        if row[0] is None:
            continue

        record = dict(zip(_COLUMNS_PAYMENTS, row))
        if (record.get("status") or "").strip().upper() == "DELETED":
            continue

        payment_type = (record.get("payment_type") or "").strip().upper()
        # RECEIPT: money comes FROM the real party (person_text) INTO cash.
        # PAYMENT: money goes OUT of cash TO the real party (target_text).
        person_text = record.get("target_text") if payment_type == "PAYMENT" else record.get("person_text")

        rows.append({
            "row_num": row_num,
            "date": (record.get("date") or "").strip(),
            "diler": (record.get("diler") or "").strip(),
            "payment_type": payment_type,
            "person_text": (person_text or "").strip(),
            "izoh": (record.get("izoh") or "").strip(),
            "amount": parse_numeric(record.get("amount")),
            # This format has no per-row account column in any real file seen
            # so far -- every active row's "Mijozga" reads "naqt pul"/"Naqd".
            "payment_account": "NAQT",
        })

    return rows


def _read_receipts_list_format(worksheet, header_row: int) -> List[Dict]:
    rows = []
    for row_num, row in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, max_col=len(_COLUMNS_RECEIPTS_LIST), values_only=True),
        start=header_row + 1,
    ):
        if row[0] is None:
            continue

        record = dict(zip(_COLUMNS_RECEIPTS_LIST, row))
        date_val = record.get("date")
        date_str = date_val.strftime("%d.%m.%Y") if hasattr(date_val, "strftime") else str(date_val or "").strip()

        rows.append({
            "row_num": row_num,
            "date": date_str,
            "diler": "",  # this format is always single-branch; Kassa Import's own `branch` field supplies it
            "payment_type": "RECEIPT",  # file name literally means "list of receipts" -- no other type seen
            "person_text": str(record.get("person_text") or "").strip(),
            "izoh": str(record.get("info") or "").strip(),
            "amount": parse_numeric(record.get("amount")),
            "payment_account": _normalize_payment_account(record.get("payment_account")),
        })

    return rows


def _find_header_row(worksheet) -> "tuple[Optional[int], Optional[str]]":
    for row_num, row in enumerate(worksheet.iter_rows(min_row=1, max_row=30), start=1):
        first = str(row[0].value or "").strip().lower()
        second = str(row[1].value or "").strip().lower() if len(row) > 1 else ""
        if first == "sana" and second == "diler":
            return row_num, "payments"
        if first in ("ид", "id") and second.startswith("плательщик"):
            return row_num, "receipts_list"
    return None, None


def extract_phone(text: str) -> Optional[str]:
    """
    Pull the phone number out of a "Name (+998 XX XXX XX XX)"-style string.

    Returns it exactly as matched (spacing/format varies in the source data --
    the caller normalizes before comparing against Customer.mobile_no).
    """
    if not text:
        return None
    match = _PHONE_RE.search(text)
    return match.group(0).strip() if match else None


def normalize_phone(phone: Optional[str]) -> Optional[str]:
    """Digits only, for tolerant comparison against however Customer.mobile_no is formatted."""
    if not phone:
        return None
    digits = re.sub(r"\D", "", phone)
    return digits or None
