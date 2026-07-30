from typing import Dict, List, Optional
from dataclasses import dataclass
from datetime import datetime, date

import frappe
from frappe import _

from akfa_diller.akfa_diller.utils.helpers import parse_numeric, get_file_path


@dataclass
class ExcelColumn:
    """Excel column configuration."""
    field_name: str
    headers: List[str]


DEFAULT_PARTY_ALIASES = ["mijoz", "mijoz nomi", "customer", "клиент", "покупатель"]
DEFAULT_PARTY_LABEL = "Mijoz"


class ExcelService:
    """
    Service for reading and parsing a daily sales/purchase Excel file.

    Expected flat-table format (one row = one item on one document):
    - Sana (MAJBURIY -- hujjat sanasi; import shaklidagi sana faqat ma'lumot
      uchun, qatorlarga o'rnini bosmaydi)
    - Mijoz/Ta'minotchi (MAJBURIY -- tizimdagi mavjud Customer/Supplier bilan
      aniq mos kelishi kerak; qaysi tomon ekani konstruktordagi
      party_aliases/party_label orqali sozlanadi)
    - Nomi (mahsulot nomi)
    - Soni (miqdori)
    - Narxi (bir dona narxi)

    Guruhlash: bitta Mijoz/Ta'minotchi + bitta Sana = bitta hujjat (alohida
    "Chek №" ustuni YO'Q -- manba ma'lumotida bunday raqam mavjud emas).

    Ikkita "party" (tomon) ishlatuvchi holat bor -- Sales uchun Mijoz/Customer,
    Purchase uchun Ta'minotchi/Supplier. Qator formatlari va parsing logikasi
    bir xil, faqat "party" ustuni sarlavhalari (aliases) va xato xabaridagi
    label farq qiladi -- shu ikkisi konstruktorda sozlanadi, qolgan hamma narsa
    umumiy.
    """

    # Rows to skip (summary/total rows) -- not party-dependent, stays static.
    SKIP_KEYWORDS = ["jami", "итого", "всего", "total", "сумма", "umumiy"]

    def __init__(self, party_aliases: Optional[List[str]] = None, party_label: str = DEFAULT_PARTY_LABEL):
        self._ensure_openpyxl()
        self.party_label = party_label
        self.COLUMNS = [
            ExcelColumn("item_name", ["nomi", "mahsulot nomi", "mahsulot", "tovar nomi", "tovar", "наименование"]),
            ExcelColumn("qty", ["soni", "miqdori", "miqdor", "количество", "количество, шт.", "кол-во"]),
            ExcelColumn("rate", ["narxi", "sotuv narxi", "narx", "цена продажи", "цена"]),
            ExcelColumn("datetime", ["sana", "savdo sanasi", "дата", "дата продажи", "datetime", "date"]),
            ExcelColumn("customer", party_aliases or DEFAULT_PARTY_ALIASES),
        ]
        self.REQUIRED_COLUMNS = ("item_name", "qty", "datetime", "customer")

    def _ensure_openpyxl(self):
        try:
            import openpyxl  # noqa
        except ImportError:
            frappe.throw(_("openpyxl o'rnatilmagan. Ishga tushiring: pip install openpyxl"))

    def read_report(self, file_url: str) -> Dict:
        """
        Read the daily sales/purchase Excel file.

        Returns:
            Dict with keys:
                - items: List of dicts with keys: item_name, qty, rate, date,
                  customer, row_num
        Raises:
            frappe.ValidationError: if the file can't be read or required columns
            are missing.
        """
        from openpyxl import load_workbook

        file_path = get_file_path(file_url)
        if not file_path:
            frappe.throw(_("Excel fayl topilmadi: {0}").format(file_url))

        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        try:
            column_indices = self._find_columns(ws)
            header_row = column_indices.get("_header_row", 1)

            self._validate_required_columns(column_indices)

            items = self._read_data_rows(ws, column_indices, header_row)

            return {"items": items}
        finally:
            wb.close()

    def _find_columns(self, worksheet) -> Dict[str, int]:
        """
        Bosh qatorni topish -- ikki bosqichda:

        1) ANIQ moslik: katakcha matni biror alias bilan TO'LIQ bir xil bo'lsa.
           Bu qisqa/umumiy aliaslar (masalan "nomi") boshqa maydonning uzunroq
           aliasi ichida ("mijoz nomi") tasodifan mos kelib qolishining oldini
           oladi.
        2) QISMAN moslik (fallback): hali topilmagan REQUIRED maydonlar uchun,
           bitta katakchada bir nechta nomzod mos kelsa -- ENG UZUN aliasga ega
           maydon tanlanadi (u aniqroq/kamroq to'qnashadi).
        """
        column_indices = {}
        rows = list(worksheet.iter_rows(min_row=1, max_row=10))

        # 1-bosqich: aniq moslik
        for row_num, row in enumerate(rows, start=1):
            for col_num, cell in enumerate(row, start=1):
                if not cell.value:
                    continue

                cell_val = str(cell.value).lower().strip()

                for column in self.COLUMNS:
                    if column.field_name in column_indices:
                        continue
                    if cell_val in column.headers:
                        column_indices[column.field_name] = col_num
                        column_indices["_header_row"] = row_num
                        break

            if all(c in column_indices for c in self.REQUIRED_COLUMNS):
                return column_indices

        # 2-bosqich: hali topilmagan maydonlar uchun qisman moslik
        for row_num, row in enumerate(rows, start=1):
            for col_num, cell in enumerate(row, start=1):
                if not cell.value:
                    continue

                cell_val = str(cell.value).lower().strip()

                best_field = None
                best_alias_len = -1
                for column in self.COLUMNS:
                    if column.field_name in column_indices:
                        continue
                    for header in column.headers:
                        if header in cell_val and len(header) > best_alias_len:
                            best_field = column.field_name
                            best_alias_len = len(header)

                if best_field:
                    column_indices[best_field] = col_num
                    column_indices.setdefault("_header_row", row_num)

            if all(c in column_indices for c in self.REQUIRED_COLUMNS):
                break

        return column_indices

    def _validate_required_columns(self, column_indices: Dict):
        labels = {
            "item_name": "Nomi",
            "qty": "Soni",
            "datetime": "Sana",
            "customer": self.party_label,
        }
        for field in self.REQUIRED_COLUMNS:
            if field not in column_indices:
                frappe.throw(_("Excel faylida '{0}' ustuni topilmadi").format(labels[field]))

    def _read_data_rows(
        self,
        worksheet,
        column_indices: Dict[str, int],
        header_row: int,
    ) -> List[Dict]:
        items = []

        item_name_col = column_indices["item_name"] - 1
        qty_col = column_indices["qty"] - 1
        dt_col = column_indices["datetime"] - 1
        customer_col = column_indices["customer"] - 1
        rate_col = column_indices.get("rate", 0) - 1 if "rate" in column_indices else None

        for row_num, row in enumerate(
            worksheet.iter_rows(min_row=header_row + 1),
            start=header_row + 1,
        ):
            item_name = self._get_cell_value(row, item_name_col)
            if not item_name:
                continue

            if self._is_summary_row(item_name):
                continue

            qty = parse_numeric(self._get_cell_value(row, qty_col))
            if qty <= 0:
                continue

            rate = 0.0
            if rate_col is not None and rate_col >= 0:
                rate = parse_numeric(self._get_cell_value(row, rate_col))

            item_date = self._parse_cell_date(row[dt_col].value)
            customer = self._get_cell_value(row, customer_col)

            items.append({
                "item_name": item_name,
                "qty": qty,
                "rate": rate,
                "date": item_date,
                "customer": customer,
                "row_num": row_num,
            })

        return items

    def _parse_cell_date(self, cell_value) -> Optional[str]:
        if cell_value is None:
            return None

        if isinstance(cell_value, datetime):
            return cell_value.strftime("%Y-%m-%d")
        if isinstance(cell_value, date):
            return cell_value.strftime("%Y-%m-%d")

        cell_str = str(cell_value).strip()
        if not cell_str:
            return None

        for fmt in (
            "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y",
            "%Y-%m-%d %H:%M:%S", "%d.%m.%Y %H:%M:%S", "%d-%m-%Y %H:%M:%S",
        ):
            try:
                return datetime.strptime(cell_str, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue

        return None

    def _get_cell_value(self, row, col_index: int) -> Optional[str]:
        if col_index < 0 or col_index >= len(row):
            return None

        value = row[col_index].value
        if value is None:
            return None

        return str(value).strip()

    def _is_summary_row(self, item_name: str) -> bool:
        # Butun (bo'sh joy/ikki nuqta olib tashlangan) katakcha kalit so'zga
        # TENG bo'lsagina jamlama qatori deb hisoblanadi -- shu kalit so'zni
        # ICHIGA olgan haqiqiy tovar nomi (masalan "Total Quartz 5W40")
        # noto'g'ri tashlab yuborilmasligi uchun substring emas, aniq moslik.
        normalized = item_name.strip().strip(":").strip().lower()
        return normalized in self.SKIP_KEYWORDS


# Sales (Mijoz/Customer) uchun standart nusxa.
excel_service = ExcelService()

# Purchase (Ta'minotchi/Supplier) uchun bir xil parsing, faqat "party"
# ustunining aliaslari va xato-xabari label'i farq qiladi.
purchase_excel_service = ExcelService(
    party_aliases=[
        "ta'minotchi", "taminotchi", "ta'minotchi nomi", "yetkazib beruvchi",
        "supplier", "поставщик",
    ],
    party_label="Ta'minotchi",
)
